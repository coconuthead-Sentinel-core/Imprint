"""imprint/traceability.py — render the requirement set as a Traceability Matrix (.xlsx).

The third view of the same requirement data (SRS = Waterfall, this = V-Model).
Columns follow the Codex Source Library template exactly:

    Requirement | Design Artifact | Implementation | Test | Evidence | Status

The requirement rows are pre-filled from the stored records; the Design/
Implementation/Test/Evidence columns start blank — you (or the assistant, later)
fill them in as the project progresses. Built on openpyxl (already installed),
UI-free so it's unit-tested without a window.
"""

from __future__ import annotations

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Exact column order from the Codex traceability-matrix-template.
COLUMNS = ["Requirement", "Design Artifact", "Implementation", "Test", "Evidence", "Status"]
_WIDTHS = [52, 22, 22, 16, 22, 14]
_HEADER_FILL = PatternFill("solid", fgColor="305496")


def default_matrix_path(project_name: str, root: str | None = None) -> str:
    """Sensible save location: <root>/output/<Project>_Traceability.xlsx."""
    root = root or os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    safe = re.sub(r"[^\w\- ]+", "", project_name).strip().replace(" ", "_") or "Project"
    return os.path.join(root, "output", f"{safe}_Traceability.xlsx")


def build_traceability_xlsx(project, requirements, out_path: str) -> str:
    """Write a traceability-matrix .xlsx for `project` + its `requirements`; return the path."""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Traceability"

    # Header row.
    ws.append(COLUMNS)
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # One row per requirement; keys traced back to the SRS.
    for q in requirements:
        ws.append([
            f"{q['req_key']}: {q['statement']}",
            "",   # Design Artifact — filled as the project designs
            "",   # Implementation — filled as code lands
            "",   # Test — filled as tests are written
            "",   # Evidence — filled at verification
            q["status"],
        ])

    # Layout: column widths, wrapped text, frozen header.
    for i, width in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return out_path
