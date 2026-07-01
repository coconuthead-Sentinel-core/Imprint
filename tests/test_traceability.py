"""Tests for the traceability-matrix renderer — writes a real .xlsx and reads it back.

Run from the project root:  py -3 -m unittest discover -s tests
"""

import os
import tempfile
import unittest

from openpyxl import load_workbook

from imprint import db, traceability


class TestTraceability(unittest.TestCase):
    def setUp(self):
        fd, self.dbpath = tempfile.mkstemp(suffix=".db")
        os.close(fd)
        self.conn = db.connect(self.dbpath)
        db.init_schema(self.conn)
        self.pid = db.create_project(self.conn, "Imprint", "vmodel", "V-Model project.")
        db.add_requirement(self.conn, self.pid, "Functional",
                           "The system shall store each requirement as a structured record.")
        db.add_requirement(self.conn, self.pid, "Non-Functional",
                           "The system shall run fully offline.")
        self.out = tempfile.mktemp(suffix=".xlsx")

    def tearDown(self):
        self.conn.close()
        os.unlink(self.dbpath)
        if os.path.exists(self.out):
            os.unlink(self.out)

    def test_headers_match_codex_template(self):
        project = db.get_project(self.conn, self.pid)
        reqs = db.list_requirements(self.conn, self.pid)
        traceability.build_traceability_xlsx(project, reqs, self.out)
        ws = load_workbook(self.out).active
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers,
                         ["Requirement", "Design Artifact", "Implementation", "Test", "Evidence", "Status"])

    def test_one_row_per_requirement(self):
        project = db.get_project(self.conn, self.pid)
        reqs = db.list_requirements(self.conn, self.pid)
        traceability.build_traceability_xlsx(project, reqs, self.out)
        ws = load_workbook(self.out).active
        self.assertEqual(ws.max_row, 1 + len(reqs))  # header + one per requirement
        first_req_cell = ws.cell(row=2, column=1).value
        self.assertIn("REQ-0001", first_req_cell)
        self.assertIn("structured record", first_req_cell)

    def test_default_path_is_xlsx_and_sanitized(self):
        p = traceability.default_matrix_path("My Project: v2!")
        self.assertTrue(p.endswith(".xlsx"))
        self.assertNotIn(":", os.path.basename(p))


if __name__ == "__main__":
    unittest.main()
