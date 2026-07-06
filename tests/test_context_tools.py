"""Tests for the assistant's context tools ported from Sentinel Forge:
doc_index (cached file indexing incl. Excel), retrieval (pure RAG), and
doc_writer (real .docx/.xlsx output with live formulas)."""
import os
import tempfile
import unittest
from datetime import datetime

from imprint import doc_index, doc_writer as dw
from imprint.retrieval import retrieve_from_index, retrieve_from_text


class DocIndexTest(unittest.TestCase):
    def _tree(self):
        d = tempfile.mkdtemp()
        os.makedirs(os.path.join(d, "Paperwork"))
        os.makedirs(os.path.join(d, ".git"))
        with open(os.path.join(d, "Paperwork", "lease.md"), "w",
                  encoding="utf-8") as f:
            f.write("the monthly rent is due on the first")
        with open(os.path.join(d, ".git", "x.md"), "w",
                  encoding="utf-8") as f:
            f.write("never index git internals")
        return d

    def test_excludes_and_relative_labels(self):
        d = self._tree()
        idx = doc_index.build_index_over(d, os.path.join(d, "c.json"))
        self.assertEqual([lbl for lbl, _ in idx],
                         [os.path.join("Paperwork", "lease.md")])

    def test_cache_reused(self):
        d = self._tree()
        cache = os.path.join(d, "c.json")
        doc_index.build_index_over(d, cache)
        idx2 = doc_index.build_index_over(d, cache)
        self.assertIn("rent", idx2[0][1])

    @unittest.skipIf(doc_index._openpyxl is None, "openpyxl not installed")
    def test_xlsx_extraction(self):
        import openpyxl
        d = tempfile.mkdtemp()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Budget"
        ws.append(["Item", "Amount"])
        ws.append(["Hosting", 60])
        p = os.path.join(d, "b.xlsx")
        wb.save(p)
        text = doc_index.extract_text(p)
        self.assertIn("Sheet: Budget", text)
        self.assertIn("Hosting | 60", text)


class RetrievalTest(unittest.TestCase):
    def test_retrieve_from_index_ranks(self):
        docs = [("a.md", "cats and dogs"),
                ("b.md", "the monthly rent is 900 dollars rent rent")]
        hits = retrieve_from_index("how much is my rent", docs)
        self.assertIn("b.md", hits)
        self.assertIn("900", hits)

    def test_retrieve_from_text_falls_back_to_opening(self):
        self.assertTrue(retrieve_from_text("zzz", "short doc").startswith(
            "short doc"))


class DocWriterTest(unittest.TestCase):
    def test_parse_table(self):
        t, h, r = dw.parse_table("Budget\nItem|Amount\nRent|$1,200")
        self.assertEqual((t, h, r), ("Budget", ["Item", "Amount"],
                                     [["Rent", 1200.0]]))

    def test_refusal_detector(self):
        self.assertTrue(dw.looks_like_refusal("I cannot do that"))
        self.assertFalse(dw.looks_like_refusal("Dear Sir,"))

    @unittest.skipIf(dw._openpyxl is None, "openpyxl not installed")
    def test_xlsx_with_sum_formula_and_safe_title(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "b.xlsx")
        dw.write_table_xlsx(p, "Budget: July", ["Item", "Amount"],
                            [["Rent", 900.0], ["Food", 250.0]])
        import openpyxl
        ws = openpyxl.load_workbook(p).active
        self.assertEqual(ws.title, "Budget July")   # ':' stripped
        self.assertEqual(ws.cell(row=4, column=2).value, "=SUM(B2:B3)")

    @unittest.skipIf(dw._docx is None, "python-docx not installed")
    def test_docx_paragraphs(self):
        d = tempfile.mkdtemp()
        p = os.path.join(d, "l.docx")
        dw.write_letter_docx(p, None, "One.\n\nTwo.")
        import docx
        self.assertEqual([x.text for x in docx.Document(p).paragraphs
                          if x.text], ["One.", "Two."])

    def test_filename_sanitized(self):
        name = dw.suggest_filename("xlsx", "a/b:c?",
                                   when=datetime(2026, 7, 6, 9, 0))
        self.assertEqual(name, "a b c 2026-07-06 0900.xlsx")


if __name__ == "__main__":
    unittest.main()
